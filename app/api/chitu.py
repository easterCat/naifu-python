import random
import time

import requests
from fake_useragent import UserAgent
from flask_restx import Namespace, reqparse, Resource
from openpyxl import load_workbook

from app import db
from app.model.template import TemplateChitu

ns = Namespace("chitu", description="词图的专属方法")
parser = reqparse.RequestParser()


@ns.route("/handle")
class HandleChitu(Resource):
    def post(self):
        total = 0
        wb = load_workbook("app/static/xlsx/NovelAI_cankaotu.xlsx")
        sheet = wb["工作表 1"]
        for row in sheet.values:
            data = {
                "name": row[5],
                "author": row[12],
                "preview": row[9],
                "prompt": row[6],
                "prompt_zh": row[7],
                "n_prompt": row[8],
                "desc": row[11],
                "model": row[13],
                "file1": row[9],
                "file2": row[9],
                "file3": row[10],
                "key_word": row[2],
                "key_word2": row[4],
            }
            tem_instance = TemplateChitu(
                name=str(data["name"]).strip(),
                author=str(data["author"]).strip(),
                preview=str(data["preview"]).strip(),
                prompt=str(data["prompt"] or "").strip(),
                prompt_zh=str(data["prompt_zh"] or "").strip(),
                n_prompt=str(data["n_prompt"]).strip(),
                desc=str(data["desc"] or "").strip(),
                model=str(data["model"]).strip(),
                file1=str(data["file1"]).strip(),
                file2=str(data["file2"]).strip(),
                file3=str(data["file3"]).strip(),
                key_word=str(data["key_word"]).strip(),
                key_word2=str(data["key_word2"]).strip(),
            )
            db.session.add(tem_instance)
            db.session.commit()
            total += 1
            print(total)
        return str(total) + "条", 200


@ns.route("/update")
class UpdateChitu(Resource):
    def post(self):
        total = 0
        wb = load_workbook("app/static/xlsx/NovelAI_cankaotu.xlsx")
        sheet = wb["工作表 1"]
        for row in sheet.values:
            data = {
                "author": row[12],
                "preview": row[9],
                "prompt": row[6],
                "prompt_zh": row[7],
                "n_prompt": row[8],
                "desc": row[11],
                "model": row[13],
            }
            total += 1
            select_one = TemplateChitu.query.filter_by(prompt=data["prompt"]).update(
                {
                    "author": data["author"],
                    "preview": data["preview"],
                    "prompt_zh": data["prompt_zh"],
                    "n_prompt": data["n_prompt"],
                    "desc": data["desc"],
                    "model": data["model"],
                }
            )
            db.session.commit()
        return str(total) + "条", 200


@ns.route("/download_img")
class DownloadChitu(Resource):
    def post(self):
        total = 0
        templates = TemplateChitu.query.all()
        options = random_header_options()
        for tem in templates:
            tem_json = tem.to_json()
            file2 = tem_json["file2"]
            file3 = tem_json["file3"]
            if file2 != "":
                file2_name = file2.split("/")[-1]
                total += 1
                download_images(file2_name, file2, options["headers"], options["proxy"])
            if file3 != "":
                file3_name = file3.split("/")[-1]
                total += 1
                download_images(file3_name, file3, options["headers"], options["proxy"])

        return str(total), 200


@ns.route("/replace_url")
class ReplaceChitu(Resource):
    def post(self):
        total = 0
        templates = TemplateChitu.query.all()
        prefix = "http://www.ptg.life/static/media/article/chi_tu/"
        for tem in templates:
            tem_json = tem.to_json()
            new_data = {
                "preview": "",
                "file1": "",
                "file2": "",
                "file3": "",
            }

            preview = tem_json["preview"]
            file1 = tem_json["file1"]
            file2 = tem_json["file2"]
            file3 = tem_json["file3"]
            if preview != "":
                new_data["preview"] = prefix + preview.split("/")[-1]
            if file1 != "":
                new_data["file1"] = prefix + file1.split("/")[-1]
            if file2 != "":
                new_data["file2"] = prefix + file2.split("/")[-1]
            if file3 != "":
                new_data["file3"] = prefix + file3.split("/")[-1]
            db.session.query(TemplateChitu).filter_by(id=tem_json["id"]).update(
                new_data
            )
            db.session.commit()
            total += 1
            print(total)
        return (
            "替换成完毕,总计处理图片"
            + str(total)
            + "张"
            + " - "
            + str(time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())),
            200,
        )


def get_proxy():
    return requests.get("http://127.0.0.1:5010/get/").json()


def delete_proxy(proxy):
    requests.get("http://127.0.0.1:5010/delete/?proxy={}".format(proxy))


def random_header_options():
    ua = UserAgent().random
    url = "https://www.ptsearch.info/home/?page="
    proxy = get_proxy().get("proxy")
    headers = {
        "User-Agent": ua,
        "Referer": "http://prompttool.com/",
    }
    proxies = {"http": "http://{}".format(proxy)}
    print("当前代理的地址 ==> ", proxy)
    options = {
        "url": url,
        "proxies": proxies,
        "headers": headers,
        "proxy": proxy,
    }
    return options


def download_images(img_name, img_url, headers, proxy):
    r = requests.get(
        img_url, headers=headers, proxies={"http": "http://{}".format(proxy)}
    )
    time.sleep(random.random() * 3)
    if r.status_code == 200:
        save_path = "app/static/media/article/chi_tu/" + img_name
        with open(save_path, "wb") as f:
            f.write(r.content)
            print("下载完成" + img_url)
            return save_path
    del r


def add(insert_data):
    me = TemplateChitu(
        name=insert_data["name"] or "暂无",
        prompt=str(insert_data["prompt"]),
        preview=insert_data["preview"],
        author=insert_data["author"],
        prompt_zh=insert_data["prompt_zh"],
        n_prompt=insert_data["n_prompt"],
        desc=insert_data["desc"],
        model=insert_data["model"],
        key_word=insert_data["key_word"],
        key_word2=insert_data["key_word2"],
        file1=insert_data["file1"],
        file2=insert_data["file2"],
        file3=insert_data["file3"],
    )
    db.session.add(me)
    db.session.commit()
